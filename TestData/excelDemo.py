import openpyxl
book = openpyxl.load_workbook("C:\\Users\\ravi.malik\\Documents\\Book1.xlsx")
sheet = book.active
testcase1_Dict={}
testcase2_Dict={}
testcase3_Dict={}# This empty dictionary declare
cell = sheet.cell(row=1, column=4)
#print(cell.value)
# sheet.cell(row=2, column=2).value = "Ravi"
# sheet.cell(row=2, column=3).value = "Malik"
# sheet.cell(row=3, column=2).value = "Pooja"
# sheet.cell(row=3, column=3).value = "Katariya"
# print(sheet.cell(row=3, column=3).value)
# print(sheet.max_row)
# print(sheet.max_column)
# for i in range(1, sheet.max_row+1):
#     #if sheet.cell(row=i, column=1).value =="testcase2":
#         for j in range(2, sheet.max_column+1):
#             testcase1_Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=2, column=j).value
#             #print(sheet.cell(row=i, column=j).value)
#             print(testcase1_Dict)

# for i in range(1, sheet.max_row+1):
#     if sheet.cell(row=i, column=1).value =="testcase1":
#         for j in range(2, sheet.max_column+1):
#             testcase1_Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=2, column=j).value
#     elif sheet.cell(row=i, column=1).value =="testcase2":
#         for j in range(2, sheet.max_column+1):
#             testcase2_Dict[sheet.cell(row=1, column=j).value]=sheet.cell(row=3, column=j).value
#     elif sheet.cell(row=i, column=1).value == "testcase3":
#         for j in range(2, sheet.max_column + 1):
#             # Dict.update(Dict)
#             testcase3_Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=4, column=j).value
# print([testcase1_Dict,testcase2_Dict,testcase3_Dict])

for i in range(2, sheet.max_row + 1):  # Start from the second row to skip headers
    # Initialize a new dictionary for each row
    row_dict = {}
    # Iterate through the columns starting from the second column
    for j in range(2, sheet.max_column + 1):
        # Use the first row for keys and the current row for values
        key = sheet.cell(row=1, column=j).value  # Column headers as keys
        value = sheet.cell(row=i, column=j).value  # Corresponding cell value
        # Add the key-value pair to the dictionary
        row_dict[key] = value
    # Print the dictionary for the current row
    print(f"Dictionary for row {i}: {row_dict}")